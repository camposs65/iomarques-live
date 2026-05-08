import json
import os
import subprocess
import sys
import tempfile
import textwrap
import tkinter as tk
import uuid
import ctypes
from collections import defaultdict
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from ctypes import wintypes

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None


COLUMNS = ("valor", "codigo", "cliente", "suplente", "tempo")
HEADERS = {
    "valor": "Valor",
    "codigo": "Código",
    "cliente": "Cliente",
    "suplente": "Suplente",
    "tempo": "Tempo",
}
COLUMN_WIDTHS = (130, 110, 250, 250, 115)
INDEX_COLUMN_WIDTH = 58

VALUE_COL = COLUMNS.index("valor")
CODE_COL = COLUMNS.index("codigo")
CLIENT_COL = COLUMNS.index("cliente")
SUPLENTE_COL = COLUMNS.index("suplente")
TIME_COL = COLUMNS.index("tempo")
CLEARABLE_CLIENT_COLS = (CLIENT_COL, SUPLENTE_COL)

if getattr(sys, "frozen", False):
    APP_DIR = Path(sys.executable).resolve().parent
    RESOURCE_DIR = Path(getattr(sys, "_MEIPASS", APP_DIR))
else:
    APP_DIR = Path(__file__).resolve().parent
    RESOURCE_DIR = APP_DIR
AUTOSAVE_PATH = APP_DIR / "live_atual.json"
HISTORY_PATH = APP_DIR / "historico_lives.json"
ASSETS_DIR = RESOURCE_DIR / "assets"
BACKUP_DIR = APP_DIR / "backups"
AUTOSAVE_BACKUP_PATH = BACKUP_DIR / "live_atual.bak.json"
HISTORY_BACKUP_PATH = BACKUP_DIR / "historico_lives.bak.json"
AUTOSAVE_INTERVAL_MS = 10_000
SNAPSHOT_INTERVAL_SECONDS = 60
MAX_LIVE_SNAPSHOTS = 30
SUMMARY_SEPARATOR = "-------------------------------------------------------------"
CHECKBOX_TEXT = "[ ]"
REPORT_TWO_COLUMN_MIN_LINES = 38
REPORT_PRINT_WIDTH_ONE_COLUMN = 90
REPORT_PRINT_WIDTH_TWO_COLUMNS = 36

COLORS = {
    "app_bg": "#F9F0F5",
    "primary": "#7B3F7A",
    "secondary": "#EDD6ED",
    "text": "#2C1040",
    "button_text": "#4A2060",
    "table_bg": "#FFFFFF",
    "table_header": "#F3F3F3",
    "table_header_text": "#2C1040",
    "grid": "#D7D0D7",
    "sold_row": "#EAF7EA",
    "active_border": "#7B3F7A",
    "timer_bg": "#FFFFFF",
}


class LiveSalesApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("IoMarques Brechó - Controle de Vendas da Live")
        self.geometry("1160x720")
        self.minsize(960, 580)

        self.row_vars = []
        self.index_frames = []
        self.index_labels = []
        self.cell_frames = []
        self.cell_entries = []
        self.clear_buttons = []
        self.active_cell = None
        self.hovered_cell = None
        self.undo_stack = []
        self.redo_stack = []
        self.restoring_rows = False
        self.max_undo_steps = 100
        self.search_var = tk.StringVar()
        self.filter_vars = [tk.StringVar() for _ in COLUMNS]
        self.filter_buttons = []

        self.live_running = False
        self.live_id = None
        self.live_first_started_at = None
        self.live_started_at = None
        self.live_finished_at = None
        self.live_elapsed_seconds = 0
        self.timer_job = None
        self.autosave_job = None
        self.last_live_snapshot_at = None
        self.logo_header_image = None
        self.logo_icon_image = None

        self._setup_style()
        self._build_ui()

        saved_data = self._read_saved_data()
        self._restore_live_state(saved_data)
        loaded = self._load_saved_rows(saved_data)
        if not loaded:
            self._add_row()

        self._read_history()
        self._refresh_totals()
        self._refresh_live_controls()
        self._schedule_timer_tick()
        self._schedule_periodic_autosave()
        self._bind_shortcuts()
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _asset_path(self, filename):
        path = ASSETS_DIR / filename
        return path if path.exists() else None

    def _load_photo(self, filename, subsample=None):
        path = self._asset_path(filename)
        if path is None:
            return None
        try:
            image = tk.PhotoImage(file=str(path))
            if subsample:
                image = image.subsample(subsample, subsample)
            return image
        except tk.TclError:
            return None

    def _setup_style(self):
        self.configure(bg=COLORS["app_bg"])

        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        style.configure(
            "Primary.TButton",
            background=COLORS["primary"],
            foreground="#FFFFFF",
            font=("Segoe UI Semibold", 10),
            padding=(14, 9),
            borderwidth=0,
            focusthickness=0,
        )
        style.map(
            "Primary.TButton",
            background=[("active", "#6C366B"), ("pressed", "#5D2F5D")],
            foreground=[("active", "#FFFFFF")],
        )
        style.configure(
            "Secondary.TButton",
            background=COLORS["secondary"],
            foreground=COLORS["button_text"],
            font=("Segoe UI Semibold", 10),
            padding=(14, 9),
            borderwidth=0,
            focusthickness=0,
        )
        style.map(
            "Secondary.TButton",
            background=[("active", "#E3C6E3"), ("pressed", "#D8B6D8")],
            foreground=[("active", COLORS["button_text"])],
        )

    def _build_ui(self):
        self.brand_header = tk.Frame(self, bg=COLORS["primary"])
        self.brand_header.pack(fill="x")

        header_inner = tk.Frame(self.brand_header, bg=COLORS["primary"])
        header_inner.pack(fill="x", padx=24, pady=18)

        self.logo_header_image = self._load_photo("logo_round.png", subsample=4)
        if self.logo_header_image is not None:
            logo_label = tk.Label(header_inner, image=self.logo_header_image, bg=COLORS["primary"], bd=0)
            logo_label.pack(side="left", padx=(0, 14))

        self.logo_icon_image = self._load_photo("app_icon.png")
        if self.logo_icon_image is not None:
            self.iconphoto(True, self.logo_icon_image)

        title_block = tk.Frame(header_inner, bg=COLORS["primary"])
        title_block.pack(side="left", fill="x", expand=True)

        title = tk.Label(
            title_block,
            text="IoMarques Brechó",
            bg=COLORS["primary"],
            fg="#FFFFFF",
            font=("Segoe UI Semibold", 24),
        )
        title.pack(anchor="w")

        subtitle = tk.Label(
            title_block,
            text="controle de vendas da live no Instagram",
            bg=COLORS["primary"],
            fg=COLORS["app_bg"],
            font=("Segoe UI", 11),
        )
        subtitle.pack(anchor="w", pady=(3, 0))

        self.toolbar = tk.Frame(self, bg=COLORS["app_bg"])
        self.toolbar.pack(fill="x", padx=24, pady=(18, 14))

        self.toolbar_actions = tk.Frame(self.toolbar, bg=COLORS["app_bg"])
        self.toolbar_actions.pack(side="left", fill="x", expand=True)

        self.pre_live_frame = tk.Frame(self.toolbar_actions, bg=COLORS["app_bg"])
        self.live_frame = tk.Frame(self.toolbar_actions, bg=COLORS["app_bg"])
        self.post_live_frame = tk.Frame(self.toolbar_actions, bg=COLORS["app_bg"])

        self.start_button = ttk.Button(
            self.pre_live_frame, text="Iniciar live", command=self.start_live, style="Primary.TButton"
        )
        self.start_button.pack(side="left", padx=(0, 8))
        self._build_final_action_buttons(self.pre_live_frame)

        self.finish_button = ttk.Button(
            self.live_frame, text="Finalizar live", command=self.finish_live, style="Primary.TButton"
        )
        self.finish_button.pack(side="left", padx=(0, 8))

        self._build_final_action_buttons(self.post_live_frame)

        timer_panel = tk.Frame(self.toolbar, bg=COLORS["app_bg"])
        timer_panel.pack(side="right")

        self.live_status_label = tk.Label(
            timer_panel,
            text="Aguardando início",
            bg=COLORS["app_bg"],
            fg=COLORS["button_text"],
            font=("Segoe UI Semibold", 10),
        )
        self.live_status_label.pack(side="top", anchor="e")

        self.timer_label = tk.Label(
            timer_panel,
            text="00:00:00",
            bg=COLORS["timer_bg"],
            fg=COLORS["text"],
            font=("Consolas", 18, "bold"),
            padx=14,
            pady=3,
        )
        self.timer_label.pack(side="top", anchor="e", pady=(3, 0))

        self.total_sold_label = tk.Label(
            timer_panel,
            text="Vendido: R$ 0,00",
            bg=COLORS["app_bg"],
            fg=COLORS["button_text"],
            font=("Segoe UI Semibold", 10),
        )
        self.total_sold_label.pack(side="top", anchor="e", pady=(4, 0))

        self.search_panel = tk.Frame(self, bg=COLORS["app_bg"])

        tk.Label(
            self.search_panel,
            text="Pesquisar",
            bg=COLORS["app_bg"],
            fg=COLORS["button_text"],
            font=("Segoe UI Semibold", 10),
        ).pack(side="left", padx=(0, 8))

        self.search_entry = tk.Entry(
            self.search_panel,
            textvariable=self.search_var,
            width=28,
            relief="solid",
            bd=1,
            fg=COLORS["text"],
            font=("Segoe UI", 10),
        )
        self.search_entry.pack(side="left", padx=(0, 8), ipady=4)
        self.search_entry.bind("<Return>", lambda _event: self._find_next_match())
        self.search_entry.bind("<Escape>", lambda _event: self._hide_search())

        ttk.Button(self.search_panel, text="Próximo", command=self._find_next_match, style="Secondary.TButton").pack(
            side="left", padx=(0, 8)
        )
        ttk.Button(self.search_panel, text="Limpar", command=self._clear_search, style="Secondary.TButton").pack(
            side="left", padx=(0, 8)
        )
        ttk.Button(self.search_panel, text="Fechar", command=self._hide_search, style="Secondary.TButton").pack(
            side="left"
        )

        self.table_shell = tk.Frame(self, bg=COLORS["grid"], bd=1)
        self.table_shell.pack(fill="both", expand=True, padx=24, pady=(0, 14))

        header_shell = tk.Frame(self.table_shell, bg=COLORS["grid"])
        header_shell.pack(fill="x")

        self.header_frame = tk.Frame(header_shell, bg=COLORS["grid"])
        self.header_frame.pack(side="left", fill="x", expand=True)
        self.header_spacer = tk.Frame(header_shell, bg=COLORS["grid"], width=17)
        self.header_spacer.pack(side="right", fill="y")
        self.header_spacer.pack_propagate(False)
        self._build_table_header()

        body_shell = tk.Frame(self.table_shell, bg=COLORS["grid"])
        body_shell.pack(fill="both", expand=True)

        self.canvas = tk.Canvas(
            body_shell,
            bg=COLORS["table_bg"],
            highlightthickness=0,
            borderwidth=0,
        )
        self.canvas.pack(side="left", fill="both", expand=True)

        scroll = ttk.Scrollbar(body_shell, orient="vertical", command=self.canvas.yview)
        scroll.pack(side="right", fill="y")
        self.header_spacer.configure(width=scroll.winfo_reqwidth())
        self.canvas.configure(yscrollcommand=scroll.set)

        self.body_frame = tk.Frame(self.canvas, bg=COLORS["grid"])
        self.body_window = self.canvas.create_window((0, 0), window=self.body_frame, anchor="nw")

        self.body_frame.bind("<Configure>", self._on_body_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)

        footer = tk.Label(
            self,
            text=(
                "Clique em Iniciar live para gravar os tempos. "
                "Ao preencher Valor ou Código, o Tempo da peça é salvo automaticamente."
            ),
            bg=COLORS["app_bg"],
            fg=COLORS["button_text"],
            font=("Segoe UI", 10),
        )
        footer.pack(fill="x", padx=24, pady=(0, 14))

    def _build_final_action_buttons(self, parent):
        ttk.Button(parent, text="Resumo final", command=self.show_summary, style="Primary.TButton").pack(
            side="left", padx=(0, 8)
        )
        ttk.Button(parent, text="Mensagens clientes", command=self.show_client_messages, style="Primary.TButton").pack(
            side="left", padx=(0, 8)
        )
        ttk.Button(parent, text="Exportar Excel", command=self.export_excel, style="Primary.TButton").pack(
            side="left", padx=(0, 8)
        )
        ttk.Button(parent, text="Imprimir resumo", command=self.print_report, style="Secondary.TButton").pack(
            side="left", padx=(0, 8)
        )
        ttk.Button(parent, text="Imprimir planilha", command=self.print_sheet, style="Secondary.TButton").pack(
            side="left", padx=(0, 8)
        )
        ttk.Button(parent, text="Imprimir não vendidas", command=self.print_unsold_pieces, style="Secondary.TButton").pack(
            side="left", padx=(0, 8)
        )
        ttk.Button(parent, text="Histórico de lives", command=self.show_history, style="Secondary.TButton").pack(
            side="left", padx=(0, 8)
        )
        ttk.Button(parent, text="Nova live / Limpar tudo", command=self.clear_all, style="Secondary.TButton").pack(
            side="left"
        )

    def _build_table_header(self):
        self.filter_buttons = []
        index_cell = tk.Frame(
            self.header_frame,
            bg=COLORS["table_header"],
            highlightbackground=COLORS["grid"],
            highlightthickness=1,
        )
        index_cell.grid(row=0, column=0, sticky="nsew")
        self.header_frame.grid_columnconfigure(0, weight=0, minsize=INDEX_COLUMN_WIDTH)
        tk.Label(
            index_cell,
            text="Peça",
            bg=COLORS["table_header"],
            fg=COLORS["table_header_text"],
            font=("Segoe UI Semibold", 10),
            anchor="center",
            padx=6,
            pady=9,
        ).pack(fill="both", expand=True)

        for col, header in enumerate(HEADERS[column] for column in COLUMNS):
            cell = tk.Frame(
                self.header_frame,
                bg=COLORS["table_header"],
                highlightbackground=COLORS["grid"],
                highlightthickness=1,
            )
            cell.grid(row=0, column=col + 1, sticky="nsew")
            self.header_frame.grid_columnconfigure(col + 1, weight=1, minsize=COLUMN_WIDTHS[col])

            label = tk.Label(
                cell,
                text=header,
                bg=COLORS["table_header"],
                fg=COLORS["table_header_text"],
                font=("Segoe UI Semibold", 11),
                anchor="w",
                padx=10,
                pady=9,
            )
            label.pack(side="left", fill="both", expand=True)

            filter_button = tk.Button(
                cell,
                text="▼",
                command=lambda c=col: self._show_filter_popup(c),
                bg=COLORS["table_header"],
                fg=COLORS["button_text"],
                activebackground=COLORS["secondary"],
                activeforeground=COLORS["button_text"],
                relief="flat",
                bd=0,
                width=2,
                padx=2,
                pady=0,
                font=("Segoe UI", 8),
                cursor="hand2",
                takefocus=False,
            )
            filter_button.pack(side="right", padx=(0, 4), pady=6)
            self.filter_buttons.append(filter_button)

    def _on_body_configure(self, _event=None):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self.canvas.itemconfigure(self.body_window, width=event.width)

    def _on_mousewheel(self, event):
        if not self.canvas.winfo_ismapped() or not self._pointer_inside_widget(self.canvas):
            return
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _bind_shortcuts(self):
        self.bind_all("<Control-z>", self.undo)
        self.bind_all("<Control-Z>", self.undo)
        self.bind_all("<Control-y>", self.redo)
        self.bind_all("<Control-Y>", self.redo)
        self.bind_all("<Control-f>", self.focus_search)
        self.bind_all("<Control-F>", self.focus_search)

    def focus_search(self, _event=None):
        if not self.search_panel.winfo_ismapped():
            self.search_panel.pack(fill="x", padx=24, pady=(0, 10), before=self.table_shell)
        self.search_entry.focus_set()
        self.search_entry.selection_range(0, tk.END)
        return "break"

    def _on_key_press(self, event, row, col):
        if not self._cell_exists(row, col) or self.restoring_rows:
            return
        if event.state & 0x4:
            return
        ignored = {
            "Shift_L",
            "Shift_R",
            "Control_L",
            "Control_R",
            "Alt_L",
            "Alt_R",
            "Caps_Lock",
            "Escape",
            "Return",
            "Tab",
            "ISO_Left_Tab",
            "Up",
            "Down",
            "Left",
            "Right",
            "Home",
            "End",
            "Prior",
            "Next",
        }
        if event.keysym in ignored:
            return
        self._record_undo_state()

    def _record_undo_state(self):
        if self.restoring_rows:
            return
        snapshot = self._rows(include_empty=True)
        if self.undo_stack and self.undo_stack[-1] == snapshot:
            return
        self.undo_stack.append(snapshot)
        if len(self.undo_stack) > self.max_undo_steps:
            self.undo_stack.pop(0)
        self.redo_stack.clear()

    def undo(self, _event=None):
        if not self.undo_stack:
            return "break"
        current = self._rows(include_empty=True)
        snapshot = self.undo_stack.pop()
        if snapshot == current and self.undo_stack:
            snapshot = self.undo_stack.pop()
        self.redo_stack.append(current)
        self._restore_rows_snapshot(snapshot)
        return "break"

    def redo(self, _event=None):
        if not self.redo_stack:
            return "break"
        current = self._rows(include_empty=True)
        snapshot = self.redo_stack.pop()
        self.undo_stack.append(current)
        self._restore_rows_snapshot(snapshot)
        return "break"

    def _restore_rows_snapshot(self, rows):
        active_cell = self.active_cell
        self.restoring_rows = True
        try:
            while len(self.row_vars) < len(rows):
                self._add_row()
            while len(self.row_vars) > len(rows):
                self._remove_last_row()
            if not rows:
                self._add_row()

            changed_rows = set()
            for row_index, row in enumerate(rows):
                for col, column in enumerate(COLUMNS):
                    value = str(row.get(column, "")).strip()
                    if self.row_vars[row_index][col].get() != value:
                        self.row_vars[row_index][col].set(value)
                        changed_rows.add(row_index)

            self._ensure_blank_row()
            for row_index in range(len(self.row_vars)):
                if row_index in changed_rows or row_index >= len(rows):
                    self._refresh_row_style(row_index)
            if active_cell and self._cell_exists(*active_cell):
                self._set_active_cell(*active_cell)
                self.cell_entries[active_cell[0]][active_cell[1]].focus_set()
            self._refresh_totals()
            self._apply_filters()
            self._save_rows()
        finally:
            self.restoring_rows = False

    def _remove_last_row(self):
        if not self.row_vars:
            return
        self.index_frames[-1].destroy()
        for cell in self.cell_frames[-1]:
            cell.destroy()
        self.index_frames.pop()
        self.index_labels.pop()
        self.row_vars.pop()
        self.cell_frames.pop()
        self.cell_entries.pop()
        self.clear_buttons.pop()
        if self.active_cell and self.active_cell[0] >= len(self.row_vars):
            self.active_cell = None

    def _show_filter_popup(self, col):
        popup = tk.Toplevel(self)
        popup.title(f"Filtro - {HEADERS[COLUMNS[col]]}")
        popup.configure(bg=COLORS["app_bg"])
        popup.resizable(False, False)
        popup.transient(self)

        tk.Label(
            popup,
            text=f"Filtrar {HEADERS[COLUMNS[col]]}",
            bg=COLORS["app_bg"],
            fg=COLORS["button_text"],
            font=("Segoe UI Semibold", 10),
        ).pack(anchor="w", padx=12, pady=(10, 4))

        entry = tk.Entry(
            popup,
            textvariable=self.filter_vars[col],
            width=28,
            relief="solid",
            bd=1,
            fg=COLORS["text"],
            font=("Segoe UI", 10),
        )
        entry.pack(fill="x", padx=12, pady=(0, 10), ipady=4)

        actions = tk.Frame(popup, bg=COLORS["app_bg"])
        actions.pack(fill="x", padx=12, pady=(0, 12))

        def apply_and_close():
            self._apply_filters()
            popup.destroy()

        def clear_and_close():
            self.filter_vars[col].set("")
            self._apply_filters()
            popup.destroy()

        ttk.Button(actions, text="Aplicar", command=apply_and_close, style="Primary.TButton").pack(
            side="left", padx=(0, 8)
        )
        ttk.Button(actions, text="Limpar", command=clear_and_close, style="Secondary.TButton").pack(side="left")

        entry.bind("<KeyRelease>", lambda _event: self._apply_filters())
        entry.bind("<Return>", lambda _event: apply_and_close())
        entry.bind("<Escape>", lambda _event: popup.destroy())

        button = self.filter_buttons[col]
        x = button.winfo_rootx()
        y = button.winfo_rooty() + button.winfo_height()
        popup.geometry(f"+{x}+{y}")
        entry.focus_set()
        entry.selection_range(0, tk.END)

    def _clear_filters(self):
        for var in self.filter_vars:
            var.set("")
        self._apply_filters()

    def _active_filters(self):
        return [var.get().strip().lower() for var in self.filter_vars]

    def _apply_filters(self):
        if not self.row_vars:
            self._refresh_filter_buttons()
            return
        filters = self._active_filters()
        has_filters = any(filters)
        for row_index, row_vars in enumerate(self.row_vars):
            visible = True
            if has_filters:
                for col, query in enumerate(filters):
                    if query and query not in row_vars[col].get().strip().lower():
                        visible = False
                        break
                if visible and not any(var.get().strip() for var in row_vars):
                    visible = False

            for col, cell in enumerate(self.cell_frames[row_index]):
                if visible:
                    self.index_frames[row_index].grid(row=row_index, column=0, sticky="nsew")
                    cell.grid(row=row_index, column=col + 1, sticky="nsew")
                else:
                    self.index_frames[row_index].grid_remove()
                    cell.grid_remove()

        self._on_body_configure()
        self._refresh_filter_buttons()

    def _refresh_filter_buttons(self):
        if not self.filter_buttons:
            return
        for col, button in enumerate(self.filter_buttons):
            active = bool(self.filter_vars[col].get().strip())
            button.configure(
                bg=COLORS["secondary"] if active else COLORS["table_header"],
                fg=COLORS["primary"] if active else COLORS["button_text"],
            )

    def _find_next_match(self):
        query = self.search_var.get().strip().lower()
        if not query:
            return

        matches = []
        for row_index, row_vars in enumerate(self.row_vars):
            for col, var in enumerate(row_vars):
                if query in var.get().strip().lower():
                    matches.append((row_index, col))

        if not matches:
            messagebox.showinfo("Pesquisa", "Não encontrei esse texto na planilha.")
            return

        start = -1
        if self.active_cell in matches:
            start = matches.index(self.active_cell)
        row, col = matches[(start + 1) % len(matches)]
        self._focus_cell(row, col)

    def _clear_search(self):
        self.search_var.set("")
        self.search_entry.focus_set()

    def _hide_search(self):
        self.search_var.set("")
        self.search_panel.pack_forget()
        if self.active_cell and self._cell_exists(*self.active_cell):
            self.cell_entries[self.active_cell[0]][self.active_cell[1]].focus_set()
        return "break"

    def _add_row(self, values=None):
        if values is None:
            values = [""] * len(COLUMNS)

        row_index = len(self.row_vars)
        row_vars = []
        row_frames = []
        row_entries = []
        row_buttons = []

        self.body_frame.grid_columnconfigure(0, weight=0, minsize=INDEX_COLUMN_WIDTH)
        index_cell = tk.Frame(
            self.body_frame,
            bg=COLORS["table_header"],
            highlightbackground=COLORS["grid"],
            highlightthickness=1,
        )
        index_cell.grid(row=row_index, column=0, sticky="nsew")
        index_label = tk.Label(
            index_cell,
            text=str(row_index + 1),
            bg=COLORS["table_header"],
            fg=COLORS["button_text"],
            font=("Segoe UI Semibold", 10),
            anchor="center",
            padx=6,
            pady=7,
        )
        index_label.pack(fill="both", expand=True)

        for col in range(len(COLUMNS)):
            self.body_frame.grid_columnconfigure(col + 1, weight=1, minsize=COLUMN_WIDTHS[col])

            cell = tk.Frame(
                self.body_frame,
                bg=COLORS["table_bg"],
                highlightbackground=COLORS["grid"],
                highlightthickness=1,
            )
            cell.grid(row=row_index, column=col + 1, sticky="nsew")

            value = values[col] if col < len(values) else ""
            var = tk.StringVar(value=value)
            entry = tk.Entry(
                cell,
                textvariable=var,
                width=1,
                relief="flat",
                bd=0,
                bg=COLORS["table_bg"],
                fg=COLORS["text"],
                insertbackground=COLORS["text"],
                font=("Segoe UI", 11),
            )
            entry.pack(side="left", fill="both", expand=True, padx=(10, 4), pady=7)

            clear_button = None
            if col in CLEARABLE_CLIENT_COLS:
                clear_button = tk.Button(
                    cell,
                    text="X",
                    command=lambda r=row_index, c=col: self._clear_client_cell(r, c),
                    bg=COLORS["table_bg"],
                    fg=COLORS["button_text"],
                    activebackground=COLORS["secondary"],
                    activeforeground=COLORS["button_text"],
                    relief="flat",
                    bd=0,
                    padx=7,
                    pady=0,
                    font=("Segoe UI Semibold", 9),
                    cursor="hand2",
                    takefocus=False,
                )

            for widget in (cell, entry, clear_button):
                if widget is None:
                    continue
                widget.bind("<Enter>", lambda _e, r=row_index, c=col: self._show_clear_button(r, c))
                widget.bind("<Leave>", lambda _e, r=row_index, c=col: self._schedule_hide_clear_button(r, c))

            entry.bind("<FocusIn>", lambda _e, r=row_index, c=col: self._set_active_cell(r, c))
            entry.bind("<FocusOut>", lambda _e, r=row_index, c=col: self._finish_cell_edit(r, c))
            entry.bind("<KeyPress>", lambda e, r=row_index, c=col: self._on_key_press(e, r, c))
            entry.bind("<KeyRelease>", lambda e, r=row_index, c=col: self._on_key_release(e, r, c))
            entry.bind("<Return>", lambda _e, r=row_index, c=col: self._move_focus(r, c, 1, 0))
            entry.bind("<Up>", lambda _e, r=row_index, c=col: self._move_focus(r, c, -1, 0))
            entry.bind("<Down>", lambda _e, r=row_index, c=col: self._move_focus(r, c, 1, 0))
            entry.bind("<Left>", lambda _e, r=row_index, c=col: self._move_focus_left(r, c))
            entry.bind("<Right>", lambda _e, r=row_index, c=col: self._move_focus_right(r, c))
            entry.bind("<Tab>", lambda _e, r=row_index, c=col: self._move_focus_right(r, c))
            entry.bind("<Shift-Tab>", lambda _e, r=row_index, c=col: self._move_focus_left(r, c))

            row_vars.append(var)
            row_frames.append(cell)
            row_entries.append(entry)
            row_buttons.append(clear_button)

        self.index_frames.append(index_cell)
        self.index_labels.append(index_label)
        self.row_vars.append(row_vars)
        self.cell_frames.append(row_frames)
        self.cell_entries.append(row_entries)
        self.clear_buttons.append(row_buttons)
        self._refresh_row_style(row_index)
        self._on_body_configure()
        if hasattr(self, "filter_vars") and any(var.get().strip() for var in self.filter_vars):
            self._apply_filters()

    def _set_active_cell(self, row, col):
        previous = self.active_cell
        self.active_cell = (row, col)
        if previous:
            self._refresh_cell_border(*previous)
        self._refresh_cell_border(row, col)

    def _refresh_cell_border(self, row, col):
        if not self._cell_exists(row, col):
            return
        color = COLORS["active_border"] if self.active_cell == (row, col) else COLORS["grid"]
        self.cell_frames[row][col].configure(highlightbackground=color)

    def _finish_cell_edit(self, row, col):
        if not self._cell_exists(row, col):
            return
        if col == VALUE_COL:
            self.row_vars[row][col].set(self._format_money_input(self.row_vars[row][col].get()))
        if col in (VALUE_COL, CODE_COL):
            self._maybe_stamp_piece_time(row)
        self._refresh_row_style(row)
        self._refresh_totals()
        self._ensure_blank_row()
        self._apply_filters()
        self._save_rows()

    def _on_key_release(self, event, row, col):
        if event.keysym in {"Up", "Down", "Left", "Right", "Return", "Tab", "ISO_Left_Tab"}:
            return
        if col in (VALUE_COL, CODE_COL):
            self._maybe_stamp_piece_time(row)
        self._refresh_row_style(row)
        self._refresh_totals()
        self._ensure_blank_row()
        self._apply_filters()
        self._save_rows()
        if self.hovered_cell == (row, col):
            self._show_clear_button(row, col)

    def _maybe_stamp_piece_time(self, row):
        if not self._cell_exists(row, TIME_COL):
            return
        has_piece_reference = (
            self.row_vars[row][VALUE_COL].get().strip() or self.row_vars[row][CODE_COL].get().strip()
        )
        if not has_piece_reference:
            self.row_vars[row][TIME_COL].set("")
        elif self.live_running and not self.row_vars[row][TIME_COL].get().strip():
            self.row_vars[row][TIME_COL].set(self._format_elapsed(self._current_elapsed_seconds()))

    def _move_focus_left(self, row, col):
        target_row = row
        target_col = col - 1
        if target_col < 0:
            target_row = max(0, row - 1)
            target_col = len(COLUMNS) - 1
        return self._focus_cell(target_row, target_col)

    def _move_focus_right(self, row, col):
        target_row = row
        target_col = col + 1
        if target_col >= len(COLUMNS):
            target_row = row + 1
            target_col = 0
        return self._focus_cell(target_row, target_col)

    def _move_focus(self, row, col, row_delta, col_delta):
        return self._focus_cell(row + row_delta, col + col_delta)

    def _focus_cell(self, row, col):
        if self.active_cell and self._cell_exists(*self.active_cell):
            self._finish_cell_edit(*self.active_cell)

        row = max(0, row)
        col = min(max(0, col), len(COLUMNS) - 1)
        while row >= len(self.row_vars):
            self._add_row()

        entry = self.cell_entries[row][col]
        self._set_active_cell(row, col)
        entry.focus_set()
        entry.icursor(tk.END)
        entry.selection_range(0, tk.END)
        self.canvas.after(10, lambda: self._scroll_cell_into_view(row))
        return "break"

    def _scroll_cell_into_view(self, row):
        if row < 0 or row >= len(self.cell_frames):
            return
        cell = self.cell_frames[row][0]
        canvas_height = max(1, self.canvas.winfo_height())
        y = cell.winfo_y()
        row_height = max(1, cell.winfo_height())
        view_top = self.canvas.canvasy(0)
        view_bottom = view_top + canvas_height
        total_height = max(1, self.body_frame.winfo_height())

        if y < view_top:
            self.canvas.yview_moveto(y / total_height)
        elif y + row_height > view_bottom:
            self.canvas.yview_moveto((y + row_height - canvas_height) / total_height)

    def _show_clear_button(self, row, col):
        self.hovered_cell = (row, col)
        if col not in CLEARABLE_CLIENT_COLS or not self._cell_exists(row, col):
            return

        button = self.clear_buttons[row][col]
        if button is None:
            return

        if self.row_vars[row][col].get().strip():
            bg = self._row_bg(row)
            button.configure(bg=bg)
            if not button.winfo_ismapped():
                button.pack(side="right", padx=(0, 5), pady=6)
        elif button.winfo_ismapped():
            button.pack_forget()

    def _schedule_hide_clear_button(self, row, col):
        self.after(60, lambda: self._hide_clear_button_if_pointer_left(row, col))

    def _hide_clear_button_if_pointer_left(self, row, col):
        if not self._cell_exists(row, col) or col not in CLEARABLE_CLIENT_COLS:
            return
        if self._pointer_inside_widget(self.cell_frames[row][col]):
            return
        button = self.clear_buttons[row][col]
        if button is not None and button.winfo_ismapped():
            button.pack_forget()
        if self.hovered_cell == (row, col):
            self.hovered_cell = None

    def _pointer_inside_widget(self, parent):
        widget = self.winfo_containing(self.winfo_pointerx(), self.winfo_pointery())
        while widget is not None:
            if widget is parent:
                return True
            widget = widget.master
        return False

    def _clear_client_cell(self, row, col):
        if not self._cell_exists(row, col) or col not in CLEARABLE_CLIENT_COLS:
            return

        self._record_undo_state()
        if col == CLIENT_COL:
            self.row_vars[row][CLIENT_COL].set(self.row_vars[row][SUPLENTE_COL].get().strip())
            self.row_vars[row][SUPLENTE_COL].set("")
        elif col == SUPLENTE_COL:
            self.row_vars[row][SUPLENTE_COL].set("")

        for clear_col in CLEARABLE_CLIENT_COLS:
            button = self.clear_buttons[row][clear_col]
            if button is not None:
                button.pack_forget()

        self._refresh_row_style(row)
        self._refresh_totals()
        self._ensure_blank_row()
        self._apply_filters()
        self._save_rows()

    def _refresh_row_style(self, row):
        if row < 0 or row >= len(self.row_vars):
            return

        bg = self._row_bg(row)
        for col in range(len(COLUMNS)):
            cell = self.cell_frames[row][col]
            entry = self.cell_entries[row][col]
            cell.configure(bg=bg)
            entry.configure(bg=bg, fg=COLORS["text"])
            button = self.clear_buttons[row][col]
            if button is not None:
                button.configure(bg=bg)
        if self.active_cell:
            self._refresh_cell_border(*self.active_cell)

    def _row_bg(self, row):
        if row < 0 or row >= len(self.row_vars):
            return COLORS["table_bg"]
        return COLORS["sold_row"] if self.row_vars[row][CLIENT_COL].get().strip() else COLORS["table_bg"]

    def _ensure_blank_row(self):
        if not self.row_vars:
            self._add_row()
            return
        if any(var.get().strip() for var in self.row_vars[-1]):
            self._add_row()

    def _cell_exists(self, row, col):
        return 0 <= row < len(self.row_vars) and 0 <= col < len(COLUMNS)

    def _rows(self, include_empty=False):
        rows = []
        for row_vars in self.row_vars:
            row = {column: row_vars[col].get().strip() for col, column in enumerate(COLUMNS)}
            if include_empty or any(row.values()):
                rows.append(row)
        return rows

    def _read_saved_data(self):
        data = self._read_json_with_backup(AUTOSAVE_PATH, AUTOSAVE_BACKUP_PATH, {})
        return data if isinstance(data, dict) else {}

    def _save_rows(self, show_error=True, force_snapshot=False):
        data = {
            "updated_at": datetime.now().isoformat(timespec="seconds"),
            "live": {
                "id": self.live_id,
                "running": self.live_running,
                "first_started_at": (
                    self.live_first_started_at.isoformat(timespec="seconds")
                    if self.live_first_started_at
                    else None
                ),
                "started_at": self.live_started_at.isoformat(timespec="seconds") if self.live_started_at else None,
                "finished_at": self.live_finished_at.isoformat(timespec="seconds") if self.live_finished_at else None,
                "elapsed_seconds": self.live_elapsed_seconds,
            },
            "rows": self._rows(),
        }
        saved = self._write_json_with_backup(
            AUTOSAVE_PATH,
            data,
            AUTOSAVE_BACKUP_PATH,
            "Erro ao salvar",
            f"Não consegui salvar a live atual:\n{{error}}",
            show_error=show_error,
        )
        if saved:
            self._write_live_snapshot(data, force=force_snapshot, show_error=show_error)

    def _read_json_with_backup(self, path, backup_path, default):
        data = self._read_json_file(path)
        if data is not None:
            try:
                self._copy_file_atomic(path, backup_path)
            except OSError:
                pass
            return data

        data = self._read_json_file(backup_path)
        if data is not None:
            return data

        return default

    def _read_json_file(self, path):
        if not path.exists():
            return None
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return None

    def _write_json_with_backup(self, path, data, backup_path, error_title, error_message, show_error=True):
        try:
            self._write_json_atomic(path, data)
            self._copy_file_atomic(path, backup_path)
            return True
        except OSError as exc:
            if show_error:
                messagebox.showerror(error_title, error_message.format(error=exc))
            return False

    def _write_json_atomic(self, path, data):
        path.parent.mkdir(parents=True, exist_ok=True)
        tmp_path = path.with_name(f".{path.stem}_{uuid.uuid4().hex}.tmp")
        try:
            with tmp_path.open("w", encoding="utf-8") as file:
                json.dump(data, file, ensure_ascii=False, indent=2)
                file.write("\n")
                file.flush()
                os.fsync(file.fileno())
            tmp_path.replace(path)
        finally:
            try:
                if tmp_path.exists():
                    tmp_path.unlink()
            except OSError:
                pass

    def _copy_file_atomic(self, source_path, target_path):
        target_path.parent.mkdir(parents=True, exist_ok=True)
        tmp_path = target_path.with_name(f".{target_path.stem}_{uuid.uuid4().hex}.tmp")
        try:
            with source_path.open("rb") as source, tmp_path.open("wb") as target:
                target.write(source.read())
                target.flush()
                os.fsync(target.fileno())
            tmp_path.replace(target_path)
        finally:
            try:
                if tmp_path.exists():
                    tmp_path.unlink()
            except OSError:
                pass

    def _write_live_snapshot(self, data, force=False, show_error=True):
        if not data.get("rows"):
            return

        now = datetime.now()
        if not force and self.last_live_snapshot_at is not None:
            elapsed = (now - self.last_live_snapshot_at).total_seconds()
            if elapsed < SNAPSHOT_INTERVAL_SECONDS:
                return

        snapshot_path = BACKUP_DIR / f"live_atual_{now.strftime('%Y%m%d_%H%M%S')}.json"
        try:
            self._write_json_atomic(snapshot_path, data)
            self.last_live_snapshot_at = now
            self._cleanup_live_snapshots()
        except OSError as exc:
            if show_error:
                messagebox.showerror("Erro no backup", f"Não consegui criar uma cópia de segurança da live:\n{exc}")

    def _cleanup_live_snapshots(self):
        snapshots = sorted(
            BACKUP_DIR.glob("live_atual_*.json"),
            key=lambda path: path.stat().st_mtime,
            reverse=True,
        )
        for snapshot in snapshots[MAX_LIVE_SNAPSHOTS:]:
            try:
                snapshot.unlink()
            except OSError:
                pass

    def _restore_live_state(self, data):
        live = data.get("live", {}) if isinstance(data, dict) else {}
        self.live_id = live.get("id") or None
        try:
            self.live_elapsed_seconds = max(0, int(live.get("elapsed_seconds", 0) or 0))
        except (TypeError, ValueError):
            self.live_elapsed_seconds = 0

        self.live_running = bool(live.get("running"))
        first_started_at = live.get("first_started_at")
        started_at = live.get("started_at")
        finished_at = live.get("finished_at")
        self.live_first_started_at = None
        self.live_started_at = None
        self.live_finished_at = None

        if first_started_at:
            try:
                self.live_first_started_at = datetime.fromisoformat(first_started_at)
            except ValueError:
                self.live_first_started_at = None

        if self.live_running and started_at:
            try:
                self.live_started_at = datetime.fromisoformat(started_at)
            except ValueError:
                self.live_running = False

        if finished_at:
            try:
                self.live_finished_at = datetime.fromisoformat(finished_at)
            except ValueError:
                self.live_finished_at = None

        if self.live_running and self.live_started_at is None:
            self.live_running = False

    def _load_saved_rows(self, data):
        rows = data.get("rows", []) if isinstance(data, dict) else []
        if not rows:
            return False

        for row in rows:
            values = [self._saved_row_value(row, col) for col in COLUMNS]
            values[VALUE_COL] = self._format_money_input(values[VALUE_COL])
            self._add_row(values)

        self._ensure_blank_row()
        return True

    def _saved_row_value(self, row, column):
        if column == "suplente":
            values = []
            for key in ("suplente", "suplente1", "suplente2"):
                value = str(row.get(key, "") or "").strip()
                if value and value not in values:
                    values.append(value)
            return " / ".join(values)
        return str(row.get(column, "")).strip()

    def _read_history(self):
        data = self._read_json_with_backup(HISTORY_PATH, HISTORY_BACKUP_PATH, [])
        if isinstance(data, list):
            return data
        return data.get("lives", []) if isinstance(data, dict) else []

    def _write_history(self, lives):
        data = {
            "updated_at": datetime.now().isoformat(timespec="seconds"),
            "lives": lives,
        }
        self._write_json_with_backup(
            HISTORY_PATH,
            data,
            HISTORY_BACKUP_PATH,
            "Erro ao salvar histórico",
            f"Não consegui salvar o histórico de lives:\n{{error}}",
        )

    def _current_live_stats(self):
        rows = self._rows()
        sold_rows = [row for row in rows if row["cliente"].strip()]
        total = sum((self._parse_money(row["valor"]) for row in sold_rows), Decimal("0"))
        clients = sorted({row["cliente"].strip() for row in sold_rows if row["cliente"].strip()}, key=str.lower)
        return {
            "pieces_count": len(rows),
            "sold_count": len(sold_rows),
            "clients_count": len(clients),
            "clients": clients,
            "total": total,
        }

    def _refresh_totals(self):
        if not hasattr(self, "total_sold_label"):
            return
        stats = self._current_live_stats()
        self.total_sold_label.configure(text=f"Vendido: {self._format_money(stats['total'])}")

    def _current_live_history_record(self, finished_at=None):
        stats = self._current_live_stats()
        finished_at = finished_at or self.live_finished_at or datetime.now()
        started_at = self.live_first_started_at or self.live_started_at or finished_at
        live_id = self.live_id or f"live_{started_at.strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:6]}"
        self.live_id = live_id
        return {
            "id": live_id,
            "started_at": started_at.isoformat(timespec="seconds"),
            "finished_at": finished_at.isoformat(timespec="seconds"),
            "duration": self._format_elapsed(self.live_elapsed_seconds),
            "pieces_count": stats["pieces_count"],
            "sold_count": stats["sold_count"],
            "clients_count": stats["clients_count"],
            "clients": stats["clients"],
            "total": self._format_money(stats["total"]),
            "rows": self._rows(),
        }

    def _upsert_current_live_history(self, finished_at=None):
        if not self._rows():
            return
        record = self._current_live_history_record(finished_at)
        lives = self._read_history()
        replaced = False
        for index, live in enumerate(lives):
            if live.get("id") == record["id"]:
                lives[index] = record
                replaced = True
                break
        if not replaced:
            lives.append(record)
        lives.sort(key=lambda live: live.get("finished_at", ""), reverse=True)
        self._write_history(lives)

    def show_history(self):
        lives = self._read_history()
        window = tk.Toplevel(self)
        window.title("Histórico de lives - IoMarques Brechó")
        window.geometry("900x520")
        window.configure(bg=COLORS["app_bg"])

        top = tk.Frame(window, bg=COLORS["primary"])
        top.pack(fill="x")
        tk.Label(
            top,
            text="Histórico de lives",
            bg=COLORS["primary"],
            fg="#FFFFFF",
            font=("Segoe UI Semibold", 17),
        ).pack(side="left", padx=18, pady=12)

        columns = ("finished_at", "duration", "pieces_count", "sold_count", "clients_count", "total")
        headings = {
            "finished_at": "Finalizada em",
            "duration": "Duração",
            "pieces_count": "Peças",
            "sold_count": "Vendidas",
            "clients_count": "Clientes",
            "total": "Total",
        }
        tree = ttk.Treeview(window, columns=columns, show="headings", selectmode="browse")
        tree.pack(fill="both", expand=True, padx=18, pady=18)

        for column in columns:
            tree.heading(column, text=headings[column])
            tree.column(column, width=130, anchor="w")

        actions = tk.Frame(window, bg=COLORS["app_bg"])
        actions.pack(fill="x", padx=18, pady=(0, 18))

        empty = tk.Label(
            window,
            text="Nenhuma live finalizada ainda.",
            bg=COLORS["app_bg"],
            fg=COLORS["text"],
            font=("Segoe UI", 11),
        )

        item_to_live = {}

        def selected_live():
            selection = tree.selection()
            if not selection:
                return None
            return item_to_live.get(selection[0])

        def refresh_empty_state():
            if lives:
                empty.place_forget()
                delete_button.state(["!disabled"])
            else:
                empty.place(relx=0.5, rely=0.55, anchor="center")
                delete_button.state(["disabled"])

        def populate_history():
            tree.delete(*tree.get_children())
            item_to_live.clear()
            for index, live in enumerate(lives):
                item_id = f"history_{index}"
                item_to_live[item_id] = live
                tree.insert(
                    "",
                    "end",
                    iid=item_id,
                    values=(
                        self._format_history_datetime(live.get("finished_at", "")),
                        live.get("duration", ""),
                        live.get("pieces_count", 0),
                        live.get("sold_count", 0),
                        live.get("clients_count", 0),
                        live.get("total", "R$ 0,00"),
                    ),
                )
            refresh_empty_state()

        def delete_selected_history():
            nonlocal lives

            live = selected_live()
            if live is None:
                messagebox.showinfo("Selecione uma live", "Escolha uma live do histórico para excluir.")
                return

            finished_at = self._format_history_datetime(live.get("finished_at", "")) or "data não informada"
            total = live.get("total", "R$ 0,00")
            answer = messagebox.askyesno(
                "Excluir live do histórico?",
                (
                    f"Excluir a live finalizada em {finished_at}, com total {total}?\n\n"
                    "Essa ação remove apenas o registro do histórico."
                ),
            )
            if not answer:
                return

            current_lives = self._read_history()
            live_id = live.get("id")
            if live_id:
                lives = [record for record in current_lives if record.get("id") != live_id]
            else:
                removed = False
                lives = []
                for record in current_lives:
                    if not removed and record == live:
                        removed = True
                        continue
                    lives.append(record)

            self._write_history(lives)
            populate_history()

        def open_selected_history():
            live = selected_live()
            if live is None:
                messagebox.showinfo("Selecione uma live", "Escolha uma live do histórico para abrir.")
                return
            if self._load_history_live_into_main_sheet(live):
                window.destroy()

        open_button = ttk.Button(
            actions,
            text="Abrir na planilha principal",
            command=open_selected_history,
            style="Primary.TButton",
        )
        open_button.pack(side="left")

        delete_button = ttk.Button(
            actions,
            text="Excluir live selecionada",
            command=delete_selected_history,
            style="Secondary.TButton",
        )
        delete_button.pack(side="right")
        tree.bind("<Double-1>", lambda _event: open_selected_history())
        tree.bind("<Return>", lambda _event: open_selected_history())
        tree.bind("<Delete>", lambda _event: delete_selected_history())

        populate_history()

    def _load_history_live_into_main_sheet(self, live):
        rows = live.get("rows") if isinstance(live, dict) else []
        if not rows:
            messagebox.showinfo(
                "Planilha indisponível",
                "Essa live foi salva no histórico antes do app guardar as linhas detalhadas.",
            )
            return False

        self._finish_active_cell()
        if self.live_running:
            messagebox.showwarning(
                "Live em andamento",
                "Finalize a live atual antes de abrir uma live do histórico na planilha principal.",
            )
            return False

        if self._rows():
            messagebox.showwarning(
                "Planilha principal ocupada",
                (
                    "A planilha principal já tem dados.\n\n"
                    "Use 'Nova live / Limpar tudo' antes de abrir uma live do histórico."
                ),
            )
            return False

        self.live_running = False
        self.live_id = live.get("id") or None
        self.live_first_started_at = self._parse_history_datetime(live.get("started_at", ""))
        self.live_started_at = None
        self.live_finished_at = self._parse_history_datetime(live.get("finished_at", ""))
        self.live_elapsed_seconds = self._history_elapsed_seconds(live)
        self.undo_stack.clear()
        self.redo_stack.clear()
        self.search_var.set("")
        self._hide_search()
        for var in self.filter_vars:
            var.set("")

        for child in self.body_frame.winfo_children():
            child.destroy()
        self.row_vars.clear()
        self.index_frames.clear()
        self.index_labels.clear()
        self.cell_frames.clear()
        self.cell_entries.clear()
        self.clear_buttons.clear()
        self.active_cell = None
        self.hovered_cell = None

        for row in rows:
            values = [self._saved_row_value(row, column) for column in COLUMNS]
            values[VALUE_COL] = self._format_money_input(values[VALUE_COL])
            self._add_row(values)

        self._ensure_blank_row()
        self._refresh_totals()
        self._refresh_live_controls()
        self._apply_filters()
        self.canvas.yview_moveto(0)
        self._save_rows(force_snapshot=True)
        return True

    def _history_elapsed_seconds(self, live):
        duration = str(live.get("duration", "") if isinstance(live, dict) else "").strip()
        parts = duration.split(":")
        if len(parts) == 3:
            try:
                hours, minutes, seconds = [int(part) for part in parts]
                return max(0, hours * 3600 + minutes * 60 + seconds)
            except ValueError:
                pass

        started_at = self._parse_history_datetime(live.get("started_at", "") if isinstance(live, dict) else "")
        finished_at = self._parse_history_datetime(live.get("finished_at", "") if isinstance(live, dict) else "")
        if started_at and finished_at:
            return max(0, int((finished_at - started_at).total_seconds()))
        return 0

    def _parse_history_datetime(self, value):
        if not value:
            return None
        try:
            return datetime.fromisoformat(value)
        except ValueError:
            return None

    def _show_history_live(self, live):
        rows = live.get("rows") if isinstance(live, dict) else []
        if not rows:
            messagebox.showinfo(
                "Planilha indisponível",
                "Essa live foi salva no histórico antes do app guardar as linhas detalhadas.",
            )
            return

        window = tk.Toplevel(self)
        window.title("Planilha da live - IoMarques Brechó")
        window.geometry("980x560")
        window.configure(bg=COLORS["app_bg"])

        top = tk.Frame(window, bg=COLORS["primary"])
        top.pack(fill="x")
        title = (
            f"Live {self._format_history_datetime(live.get('finished_at', ''))}"
            f" | {live.get('total', 'R$ 0,00')}"
        )
        tk.Label(
            top,
            text=title,
            bg=COLORS["primary"],
            fg="#FFFFFF",
            font=("Segoe UI Semibold", 16),
        ).pack(side="left", padx=18, pady=12)

        columns = ("peca",) + COLUMNS
        tree = ttk.Treeview(window, columns=columns, show="headings")
        tree.pack(fill="both", expand=True, padx=18, pady=18)

        headings = {"peca": "Peça", **HEADERS}
        widths = {"peca": 65, "valor": 120, "codigo": 110, "cliente": 230, "suplente": 220, "tempo": 110}
        for column in columns:
            tree.heading(column, text=headings[column])
            tree.column(column, width=widths.get(column, 130), anchor="w")

        for index, row in enumerate(rows, start=1):
            tree.insert(
                "",
                "end",
                values=(
                    index,
                    self._saved_row_value(row, "valor"),
                    self._saved_row_value(row, "codigo"),
                    self._saved_row_value(row, "cliente"),
                    self._saved_row_value(row, "suplente"),
                    self._saved_row_value(row, "tempo"),
                ),
            )

    def _format_history_datetime(self, value):
        parsed = self._parse_history_datetime(value)
        return parsed.strftime("%d/%m/%Y %H:%M") if parsed else value

    def start_live(self):
        if self.live_running:
            return
        now = datetime.now()
        if not self.live_id:
            self.live_id = f"live_{now.strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:6]}"
        if self.live_first_started_at is None:
            self.live_first_started_at = now
        self.live_started_at = now
        self.live_finished_at = None
        self.live_running = True
        self._refresh_live_controls()
        self._save_rows()
        self._schedule_timer_tick()

    def finish_live(self):
        if not self.live_running:
            return
        answer = messagebox.askyesno(
            "Finalizar live?",
            "Tem certeza que deseja finalizar a live agora?",
        )
        if not answer:
            return
        finished_at = datetime.now()
        self.live_elapsed_seconds = self._current_elapsed_seconds()
        self.live_started_at = None
        self.live_finished_at = finished_at
        self.live_running = False
        if self.timer_job is not None:
            self.after_cancel(self.timer_job)
            self.timer_job = None
        self._refresh_live_controls()
        self._save_rows(force_snapshot=True)
        self._upsert_current_live_history(finished_at)

    def _refresh_live_controls(self):
        elapsed = self._current_elapsed_seconds()
        self.timer_label.configure(text=self._format_elapsed(elapsed))
        self.pre_live_frame.pack_forget()
        self.live_frame.pack_forget()
        self.post_live_frame.pack_forget()
        if self.live_running:
            self.brand_header.pack_forget()
            self.live_status_label.configure(text="Live em andamento")
            self.live_frame.pack(side="left", fill="x")
            self.start_button.state(["disabled"])
            self.finish_button.state(["!disabled"])
        elif elapsed > 0 or self.live_finished_at is not None:
            if not self.brand_header.winfo_ismapped():
                self.brand_header.pack(fill="x", before=self.toolbar)
            self.live_status_label.configure(text="Live finalizada")
            self.post_live_frame.pack(side="left", fill="x")
            self.start_button.state(["!disabled"])
            self.finish_button.state(["disabled"])
        else:
            if not self.brand_header.winfo_ismapped():
                self.brand_header.pack(fill="x", before=self.toolbar)
            self.live_status_label.configure(text="Aguardando início")
            self.pre_live_frame.pack(side="left", fill="x")
            self.start_button.state(["!disabled"])
            self.finish_button.state(["disabled"])

    def _current_live_status(self):
        if self.live_running:
            return "Em andamento"
        if self.live_elapsed_seconds > 0:
            return "Finalizada"
        return "Não iniciada"

    def _schedule_timer_tick(self):
        self._refresh_live_controls()
        if self.live_running:
            self.timer_job = self.after(1000, self._schedule_timer_tick)
        else:
            self.timer_job = None

    def _schedule_periodic_autosave(self):
        self._save_rows(show_error=False)
        self.autosave_job = self.after(AUTOSAVE_INTERVAL_MS, self._schedule_periodic_autosave)

    def _current_elapsed_seconds(self):
        elapsed = self.live_elapsed_seconds
        if self.live_running and self.live_started_at is not None:
            elapsed += max(0, int((datetime.now() - self.live_started_at).total_seconds()))
        return elapsed

    def _format_elapsed(self, seconds):
        seconds = max(0, int(seconds))
        hours, remainder = divmod(seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

    def _parse_money_optional(self, value):
        clean = value.strip().replace("R$", "").replace("r$", "").replace(" ", "")
        if not clean:
            return None

        if "," in clean:
            clean = clean.replace(".", "").replace(",", ".")
        elif "." in clean:
            parts = clean.split(".")
            if len(parts) == 2 and 1 <= len(parts[1]) <= 2:
                clean = ".".join(parts)
            else:
                clean = clean.replace(".", "")

        try:
            return Decimal(clean)
        except InvalidOperation:
            return None

    def _parse_money(self, value):
        amount = self._parse_money_optional(value)
        return amount if amount is not None else Decimal("0")

    def _format_money_input(self, value):
        if not value.strip():
            return ""
        amount = self._parse_money_optional(value)
        if amount is None:
            return value.strip()
        return self._format_money(amount)

    def _format_money(self, value):
        quantized = Decimal(value).quantize(Decimal("0.01"))
        return f"R$ {quantized:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    def _summary(self):
        grouped = defaultdict(list)
        for row in self._rows():
            cliente = row["cliente"].strip()
            if not cliente:
                continue
            valor = self._parse_money(row["valor"])
            grouped[cliente].append(
                {
                    "codigo": row["codigo"],
                    "tempo": row["tempo"],
                    "valor_texto": row["valor"],
                    "valor": valor,
                    "suplente": row["suplente"],
                }
            )
        return dict(sorted(grouped.items(), key=lambda item: item[0].lower()))

    def _summary_text(self):
        sections = self._summary_sections()
        if not sections:
            return "Nenhuma peça com cliente titular ainda.\n"

        lines = []
        for index, section in enumerate(sections):
            if index > 0:
                lines.append(SUMMARY_SEPARATOR)
            lines.append(section["header"])
            lines.extend(section["items"])
        lines.append(SUMMARY_SEPARATOR)
        lines.extend(self._summary_totals_lines())

        return "\n".join(lines) + "\n"

    def _summary_totals_lines(self):
        stats = self._current_live_stats()
        return [
            f"Clientes: {stats['clients_count']}",
            f"Peças vendidas: {stats['sold_count']}",
            f"Total vendido: {self._format_money(stats['total'])}",
        ]

    def _summary_sections(self):
        sections = []
        for cliente, items in self._summary().items():
            subtotal = sum((item["valor"] for item in items), Decimal("0"))
            codes = ", ".join(item["codigo"] or "sem código" for item in items)
            item_lines = []
            for item in items:
                codigo = item["codigo"] or "sem código"
                tempo = item["tempo"] or "-"
                valor = self._format_money(item["valor"])
                suplente = item.get("suplente", "").strip()
                suplente_text = f"   sup: {suplente}" if suplente else ""
                item_lines.append(f"{CHECKBOX_TEXT} {codigo} - {tempo} - {valor}{suplente_text}")
            sections.append(
                {
                    "cliente": cliente,
                    "header": f"{cliente} - {codes} | Total = {self._format_money(subtotal)}",
                    "items": item_lines,
                }
            )
        return sections

    def _alternates_text(self):
        rows = [row for row in self._rows() if row.get("suplente", "").strip()]
        if not rows:
            return "Nenhuma suplente registrada.\n"

        lines = []
        for row in rows:
            codigo = row["codigo"] or "sem código"
            tempo = row["tempo"] or "-"
            valor = self._format_money(self._parse_money(row["valor"]))
            cliente = row["cliente"] or "sem cliente"
            suplente = row["suplente"]
            lines.append(f"{suplente} - peça {codigo} - {valor}")
            lines.append(f"  Titular: {cliente} | Tempo: {tempo}")
            lines.append("")
        return "\n".join(lines) + "\n"

    def _client_messages_text(self):
        summary = self._summary()
        if not summary:
            return "Nenhuma peça com cliente titular ainda.\n"

        messages = []
        for cliente, items in summary.items():
            subtotal = sum((item["valor"] for item in items), Decimal("0"))
            lines = [f"Oi, {cliente}! Suas peças da live:"]
            for item in items:
                codigo = item["codigo"] or "sem código"
                tempo = f" - {item['tempo']}" if item["tempo"] else ""
                lines.append(f"- Peça {codigo}{tempo}: {self._format_money(item['valor'])}")
            lines.append(f"Total: {self._format_money(subtotal)}")
            messages.append("\n".join(lines))
        return "\n\n---\n\n".join(messages) + "\n"

    def _clip_text(self, value, width):
        text = str(value).strip()
        if len(text) <= width:
            return text
        if width <= 3:
            return text[:width]
        return text[: width - 3] + "..."

    def _printable_report(self):
        return "\n".join(line for line, _bold in self._printable_report_lines()) + "\n"

    def _printable_report_lines(self):
        now = datetime.now().strftime("%d/%m/%Y %H:%M")
        sections = self._summary_sections()
        lines = [
            (f"IoMarques Brechó - Resumo | {now}", True),
            (f"Duração: {self._format_elapsed(self._current_elapsed_seconds())}", False),
            ("", False),
        ]
        if not sections:
            lines.append(("Nenhuma peça com cliente titular ainda.", False))
            return lines

        for index, section in enumerate(sections):
            if index > 0:
                lines.append((SUMMARY_SEPARATOR, False))
            lines.append((section["header"], True))
            lines.extend((item, False) for item in section["items"])
        lines.append((SUMMARY_SEPARATOR, False))
        lines.extend((line, True) for line in self._summary_totals_lines())
        return lines

    def _wrap_report_lines_for_print(self, print_lines, columns):
        width = REPORT_PRINT_WIDTH_TWO_COLUMNS if columns > 1 else REPORT_PRINT_WIDTH_ONE_COLUMN
        wrapped = []
        for line, bold in print_lines:
            if not line:
                wrapped.append(("", bold))
                continue
            if line == SUMMARY_SEPARATOR:
                wrapped.append(("-" * width, bold))
                continue
            parts = textwrap.wrap(
                line,
                width=width,
                subsequent_indent="  ",
                break_long_words=True,
                break_on_hyphens=False,
            )
            wrapped.extend((part, bold) for part in parts)
        return wrapped

    def show_summary(self):
        self._finish_active_cell()

        window = tk.Toplevel(self)
        window.title("Resumo final - IoMarques Brechó")
        window.geometry("760x540")
        window.configure(bg=COLORS["app_bg"])

        top = tk.Frame(window, bg=COLORS["primary"])
        top.pack(fill="x")
        tk.Label(
            top,
            text="Resumo final",
            bg=COLORS["primary"],
            fg="#FFFFFF",
            font=("Segoe UI Semibold", 17),
        ).pack(side="left", padx=18, pady=12)
        ttk.Button(top, text="Imprimir", command=self.print_report, style="Secondary.TButton").pack(
            side="right", padx=18, pady=10
        )

        text = tk.Text(
            window,
            wrap="word",
            bg="#FFFFFF",
            fg=COLORS["text"],
            relief="flat",
            font=("Consolas", 11),
            padx=16,
            pady=14,
        )
        text.pack(fill="both", expand=True, padx=18, pady=18)
        text.tag_configure("client_header", font=("Consolas", 11, "bold"))
        text.tag_configure("summary_total", font=("Consolas", 11, "bold"))
        self._insert_summary_text(text)
        text.configure(state="disabled")

    def _insert_summary_text(self, text):
        sections = self._summary_sections()
        if not sections:
            text.insert("end", "Nenhuma peça com cliente titular ainda.\n")
            return

        for index, section in enumerate(sections):
            if index > 0:
                text.insert("end", SUMMARY_SEPARATOR + "\n")
            cliente = section.get("cliente", "")
            header = section["header"]
            if cliente and header.startswith(cliente):
                text.insert("end", cliente, "client_header")
                text.insert("end", header[len(cliente) :] + "\n")
            else:
                text.insert("end", header + "\n", "client_header")
            for line in section["items"]:
                text.insert("end", line + "\n")
        text.insert("end", SUMMARY_SEPARATOR + "\n")
        for line in self._summary_totals_lines():
            text.insert("end", line + "\n", "summary_total")

    def show_client_messages(self):
        self._finish_active_cell()

        window = tk.Toplevel(self)
        window.title("Mensagens para clientes - IoMarques Brechó")
        window.geometry("760x540")
        window.configure(bg=COLORS["app_bg"])

        top = tk.Frame(window, bg=COLORS["primary"])
        top.pack(fill="x")
        tk.Label(
            top,
            text="Mensagens para clientes",
            bg=COLORS["primary"],
            fg="#FFFFFF",
            font=("Segoe UI Semibold", 17),
        ).pack(side="left", padx=18, pady=12)

        text = tk.Text(
            window,
            wrap="word",
            bg="#FFFFFF",
            fg=COLORS["text"],
            relief="flat",
            font=("Segoe UI", 11),
            padx=16,
            pady=14,
        )
        text.pack(fill="both", expand=True, padx=18, pady=(18, 10))
        text.insert("end", self._client_messages_text())

        actions = tk.Frame(window, bg=COLORS["app_bg"])
        actions.pack(fill="x", padx=18, pady=(0, 18))

        def copy_messages():
            content = text.get("1.0", "end").strip()
            self.clipboard_clear()
            self.clipboard_append(content)
            messagebox.showinfo("Copiado", "As mensagens foram copiadas.")

        ttk.Button(actions, text="Copiar mensagens", command=copy_messages, style="Primary.TButton").pack(side="right")

    def export_excel(self):
        self._finish_active_cell()

        if Workbook is None:
            messagebox.showerror(
                "Biblioteca ausente",
                "Instale a biblioteca openpyxl com:\n\npip install openpyxl",
            )
            return

        rows = self._rows()
        if not rows:
            messagebox.showinfo("Nada para exportar", "Preencha pelo menos uma linha antes de exportar.")
            return

        default_name = f"live_iomarques_brecho_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
        path = filedialog.asksaveasfilename(
            title="Salvar planilha da live",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Planilha Excel", "*.xlsx")],
        )
        if not path:
            return

        wb = Workbook()
        sales = wb.active
        sales.title = "Vendas"
        sales.append(["Peça"] + [HEADERS[col] for col in COLUMNS])
        for index, row in enumerate(rows, start=1):
            sales.append([index] + [row[col] for col in COLUMNS])

        summary_sheet = wb.create_sheet("Resumo por cliente")
        summary_sheet.append(["Peça", "Cliente", "Código", "Tempo", "Valor", "Total da cliente"])
        summary = self._summary()
        grand_total = Decimal("0")
        piece_index = 1
        for cliente, items in summary.items():
            subtotal = sum((item["valor"] for item in items), Decimal("0"))
            grand_total += subtotal
            first_row = True
            for item in items:
                summary_sheet.append(
                    [
                        piece_index,
                        cliente if first_row else "",
                        item["codigo"],
                        item["tempo"],
                        float(item["valor"]),
                        float(subtotal) if first_row else "",
                    ]
                )
                piece_index += 1
                first_row = False
        summary_sheet.append([])
        summary_sheet.append(["", "TOTAL GERAL", "", "", "", float(grand_total)])

        alternates_sheet = wb.create_sheet("Suplentes")
        alternates_sheet.append(["Peça", "Suplente", "Código", "Tempo", "Valor", "Cliente titular"])
        for index, row in enumerate((row for row in rows if row["suplente"].strip()), start=1):
            alternates_sheet.append(
                [
                    index,
                    row["suplente"],
                    row["codigo"],
                    row["tempo"],
                    float(self._parse_money(row["valor"])),
                    row["cliente"],
                ]
            )

        metadata_sheet = wb.create_sheet("Dados da live")
        stats = self._current_live_stats()
        metadata_sheet.append(["Campo", "Valor"])
        metadata_sheet.append(["ID da live", self.live_id or ""])
        metadata_sheet.append(["Status", self._current_live_status()])
        metadata_sheet.append(
            [
                "Início",
                self.live_first_started_at.strftime("%d/%m/%Y %H:%M:%S") if self.live_first_started_at else "",
            ]
        )
        metadata_sheet.append(
            [
                "Finalização",
                self.live_finished_at.strftime("%d/%m/%Y %H:%M:%S") if self.live_finished_at else "",
            ]
        )
        metadata_sheet.append(["Duração registrada", self._format_elapsed(self._current_elapsed_seconds())])
        metadata_sheet.append(["Peças preenchidas", stats["pieces_count"]])
        metadata_sheet.append(["Peças vendidas", stats["sold_count"]])
        metadata_sheet.append(["Clientes diferentes", stats["clients_count"]])
        metadata_sheet.append(["Total vendido", self._format_money(stats["total"])])

        history_sheet = wb.create_sheet("Histórico de lives")
        history_sheet.append(
            ["Finalizada em", "Iniciada em", "Duração", "Peças", "Vendidas", "Clientes", "Total", "Nomes das clientes"]
        )
        for live in self._read_history():
            history_sheet.append(
                [
                    self._format_history_datetime(live.get("finished_at", "")),
                    self._format_history_datetime(live.get("started_at", "")),
                    live.get("duration", ""),
                    live.get("pieces_count", 0),
                    live.get("sold_count", 0),
                    live.get("clients_count", 0),
                    live.get("total", "R$ 0,00"),
                    ", ".join(live.get("clients", [])),
                ]
            )

        self._format_workbook(sales, summary_sheet, alternates_sheet, metadata_sheet, history_sheet)
        wb.save(path)
        self._save_rows()
        messagebox.showinfo("Exportação concluída", f"Planilha salva em:\n{path}")

    def _format_workbook(self, *sheets):
        header_fill = PatternFill("solid", fgColor=COLORS["table_header"].replace("#", ""))
        sold_fill = PatternFill("solid", fgColor=COLORS["sold_row"].replace("#", ""))
        header_font = Font(color=COLORS["table_header_text"].replace("#", ""), bold=True)
        text_font = Font(color=COLORS["text"].replace("#", ""))

        for sheet in sheets:
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.font = text_font

            for column_cells in sheet.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(length + 3, 14), 34)

            sheet.freeze_panes = "A2"
            if sheet.max_row > 1 and sheet.max_column > 1:
                sheet.auto_filter.ref = sheet.dimensions

        for row in range(2, sheets[0].max_row + 1):
            cliente = sheets[0].cell(row=row, column=CLIENT_COL + 2).value
            if cliente:
                for col in range(1, len(COLUMNS) + 2):
                    sheets[0].cell(row=row, column=col).fill = sold_fill

        for row in range(2, sheets[1].max_row + 1):
            sheets[1].cell(row=row, column=5).number_format = '"R$" #,##0.00'
            sheets[1].cell(row=row, column=6).number_format = '"R$" #,##0.00'

        for row in range(2, sheets[2].max_row + 1):
            sheets[2].cell(row=row, column=5).number_format = '"R$" #,##0.00'

    def print_sheet(self):
        self._finish_active_cell()
        rows = self._rows()
        if not rows:
            messagebox.showinfo("Nada para imprimir", "Preencha pelo menos uma linha antes de imprimir.")
            return
        self._print_rows_as_workbook(rows, "Planilha da live", "iomarques_planilha")

    def print_unsold_pieces(self):
        self._finish_active_cell()
        rows = [
            row
            for row in self._rows()
            if not row["cliente"].strip() and (row["valor"].strip() or row["codigo"].strip())
        ]
        if not rows:
            messagebox.showinfo("Sem peças não vendidas", "Não encontrei peças preenchidas sem cliente.")
            return
        self._print_rows_as_workbook(rows, "Peças não vendidas", "iomarques_nao_vendidas")

    def _print_rows_as_workbook(self, rows, title, filename_prefix):
        printable_text = self._printable_rows_text(rows, title)
        try:
            self._send_text_to_printer(printable_text, title)
        except Exception as exc:
            path = Path(tempfile.gettempdir()) / f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            try:
                path.write_text(printable_text, encoding="utf-8")
                saved_message = f"Salvei uma cópia para impressão em:\n{path}\n\n"
            except OSError:
                saved_message = ""
            messagebox.showerror(
                "Não consegui imprimir",
                f"{saved_message}Erro ao enviar para a impressora:\n{exc}",
            )
            return

        messagebox.showinfo("Impressão enviada", "Enviei a planilha para a impressora padrão.")

    def _printable_rows_text(self, rows, title):
        now = datetime.now().strftime("%d/%m/%Y %H:%M")
        widths = {
            "check": 3,
            "peca": 4,
            "valor": 11,
            "codigo": 8,
            "cliente": 20,
            "suplente": 16,
            "tempo": 8,
        }

        def cell(value, width, align="left"):
            text = self._clip_text(value, width)
            return text.rjust(width) if align == "right" else text.ljust(width)

        header = (
            f"{cell('', widths['check'])} "
            f"{cell('Peça', widths['peca'])} "
            f"{cell('Valor', widths['valor'])} "
            f"{cell('Código', widths['codigo'])} "
            f"{cell('Cliente', widths['cliente'])} "
            f"{cell('Suplente', widths['suplente'])} "
            f"{cell('Tempo', widths['tempo'])}"
        )
        separator = "-" * len(header)
        lines = [
            f"IoMarques Brechó - {title}",
            f"Gerado em: {now} | Duração: {self._format_elapsed(self._current_elapsed_seconds())} | Linhas: {len(rows)}",
            "",
            header,
            separator,
        ]
        for index, row in enumerate(rows, start=1):
            lines.append(
                f"{cell(CHECKBOX_TEXT, widths['check'])} "
                f"{cell(index, widths['peca'], 'right')} "
                f"{cell(row['valor'], widths['valor'])} "
                f"{cell(row['codigo'], widths['codigo'])} "
                f"{cell(row['cliente'], widths['cliente'])} "
                f"{cell(row['suplente'], widths['suplente'])} "
                f"{cell(row['tempo'], widths['tempo'])}"
            )
        return "\n".join(lines) + "\n"

    def _send_text_to_printer(self, text, title):
        if sys.platform.startswith("win"):
            self._send_text_to_windows_printer(text, title)
            return

        path = Path(tempfile.gettempdir()) / f"iomarques_print_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        path.write_text(text, encoding="utf-8")
        subprocess.run(["lp", str(path)], check=True)

    def _send_text_to_windows_printer(self, text, title):
        print_lines = [(line, False) for line in (text.splitlines() or [""])]
        self._send_print_lines_to_windows_printer(print_lines, title)

    def _send_print_lines_to_windows_printer(self, print_lines, title, columns=1, paginate=False):
        printer_name = self._default_windows_printer()
        gdi32 = ctypes.WinDLL("gdi32", use_last_error=True)

        class DOCINFOW(ctypes.Structure):
            _fields_ = [
                ("cbSize", ctypes.c_int),
                ("lpszDocName", wintypes.LPCWSTR),
                ("lpszOutput", wintypes.LPCWSTR),
                ("lpszDatatype", wintypes.LPCWSTR),
                ("fwType", wintypes.DWORD),
            ]

        gdi32.CreateDCW.argtypes = [wintypes.LPCWSTR, wintypes.LPCWSTR, wintypes.LPCWSTR, ctypes.c_void_p]
        gdi32.CreateDCW.restype = wintypes.HDC
        gdi32.DeleteDC.argtypes = [wintypes.HDC]
        gdi32.GetDeviceCaps.argtypes = [wintypes.HDC, ctypes.c_int]
        gdi32.GetDeviceCaps.restype = ctypes.c_int
        gdi32.CreateFontW.argtypes = [
            ctypes.c_int,
            ctypes.c_int,
            ctypes.c_int,
            ctypes.c_int,
            ctypes.c_int,
            ctypes.c_ulong,
            ctypes.c_ulong,
            ctypes.c_ulong,
            ctypes.c_ulong,
            ctypes.c_ulong,
            ctypes.c_ulong,
            ctypes.c_ulong,
            ctypes.c_ulong,
            wintypes.LPCWSTR,
        ]
        gdi32.CreateFontW.restype = wintypes.HFONT
        gdi32.SelectObject.argtypes = [wintypes.HDC, wintypes.HGDIOBJ]
        gdi32.SelectObject.restype = wintypes.HGDIOBJ
        gdi32.DeleteObject.argtypes = [wintypes.HGDIOBJ]
        gdi32.SetBkMode.argtypes = [wintypes.HDC, ctypes.c_int]
        gdi32.StartDocW.argtypes = [wintypes.HDC, ctypes.POINTER(DOCINFOW)]
        gdi32.StartDocW.restype = ctypes.c_int
        gdi32.EndDoc.argtypes = [wintypes.HDC]
        gdi32.AbortDoc.argtypes = [wintypes.HDC]
        gdi32.StartPage.argtypes = [wintypes.HDC]
        gdi32.EndPage.argtypes = [wintypes.HDC]
        gdi32.TextOutW.argtypes = [wintypes.HDC, ctypes.c_int, ctypes.c_int, wintypes.LPCWSTR, ctypes.c_int]
        gdi32.GetTextExtentPoint32W.argtypes = [
            wintypes.HDC,
            wintypes.LPCWSTR,
            ctypes.c_int,
            ctypes.POINTER(wintypes.SIZE),
        ]
        gdi32.GetTextExtentPoint32W.restype = wintypes.BOOL

        hdc = gdi32.CreateDCW("WINSPOOL", printer_name, None, None)
        if not hdc:
            raise ctypes.WinError(ctypes.get_last_error())

        print_lines = [(str(line), bool(bold)) for line, bold in print_lines] or [("", False)]
        columns = max(1, min(2, int(columns or 1)))
        print_columns = self._split_print_columns(print_lines, columns)
        fit_lines = [line for column in print_columns for line in column]
        max_column_lines = max((len(column) for column in print_columns), default=1)
        regular_font = None
        bold_font = None
        old_font = None
        doc_started = False
        page_started = False
        try:
            dpi_y = max(1, gdi32.GetDeviceCaps(hdc, 90))
            page_width = max(1, gdi32.GetDeviceCaps(hdc, 8))
            page_height = max(1, gdi32.GetDeviceCaps(hdc, 10))
            dpi_x = max(1, gdi32.GetDeviceCaps(hdc, 88))
            margin_x = max(30, int(dpi_x * 0.18))
            margin_y = max(30, int(dpi_y * 0.18))
            printable_width = max(1, page_width - margin_x * 2)
            printable_height = max(1, page_height - margin_y * 2)
            column_gap = max(20, int(dpi_x * 0.16)) if columns > 1 else 0
            column_width = max(1, int((printable_width - column_gap * (columns - 1)) / columns))

            regular_font, bold_font, line_height = self._fit_print_font(
                gdi32,
                hdc,
                fit_lines,
                dpi_y,
                column_width,
                printable_height,
                max_lines_per_column=max_column_lines,
                fit_width=(columns == 1),
                fit_height=not paginate,
            )
            old_font = gdi32.SelectObject(hdc, regular_font)
            gdi32.SetBkMode(hdc, 1)

            docinfo = DOCINFOW(ctypes.sizeof(DOCINFOW), f"IoMarques Brechó - {title}", None, None, 0)
            if gdi32.StartDocW(hdc, ctypes.byref(docinfo)) <= 0:
                raise ctypes.WinError(ctypes.get_last_error())
            doc_started = True
            if gdi32.StartPage(hdc) <= 0:
                raise ctypes.WinError(ctypes.get_last_error())
            page_started = True

            for column_index, column in enumerate(print_columns):
                x = margin_x + column_index * (column_width + column_gap)
                y = margin_y
                printed_on_page = False
                for line, bold in column:
                    if paginate and printed_on_page and y + line_height > margin_y + printable_height:
                        if gdi32.EndPage(hdc) <= 0:
                            raise ctypes.WinError(ctypes.get_last_error())
                        if gdi32.StartPage(hdc) <= 0:
                            raise ctypes.WinError(ctypes.get_last_error())
                        y = margin_y
                        printed_on_page = False
                    selected_font = bold_font if bold else regular_font
                    gdi32.SelectObject(hdc, selected_font)
                    printable_line = self._clip_line_for_print_width(gdi32, hdc, line, column_width)
                    gdi32.TextOutW(hdc, x, y, printable_line, len(printable_line))
                    y += line_height
                    printed_on_page = True

            if gdi32.EndPage(hdc) <= 0:
                raise ctypes.WinError(ctypes.get_last_error())
            page_started = False
            if gdi32.EndDoc(hdc) <= 0:
                raise ctypes.WinError(ctypes.get_last_error())
            doc_started = False
        except Exception:
            if page_started or doc_started:
                gdi32.AbortDoc(hdc)
            raise
        finally:
            if old_font:
                gdi32.SelectObject(hdc, old_font)
            if regular_font:
                gdi32.DeleteObject(regular_font)
            if bold_font:
                gdi32.DeleteObject(bold_font)
            gdi32.DeleteDC(hdc)

    def _split_print_columns(self, print_lines, columns):
        if columns <= 1 or len(print_lines) <= REPORT_TWO_COLUMN_MIN_LINES:
            return [print_lines]
        midpoint = (len(print_lines) + 1) // 2
        separator_indexes = [
            index
            for index, (line, _bold) in enumerate(print_lines)
            if line and set(line) == {"-"} and 4 < index < len(print_lines) - 4
        ]
        if separator_indexes:
            midpoint = min(separator_indexes, key=lambda index: abs(index - midpoint)) + 1
        return [print_lines[:midpoint], print_lines[midpoint:]]

    def _fit_print_font(
        self,
        gdi32,
        hdc,
        print_lines,
        dpi_y,
        printable_width,
        printable_height,
        max_lines_per_column=None,
        fit_width=True,
        fit_height=True,
    ):
        max_lines = max(1, max_lines_per_column or len(print_lines))
        height_points = int(printable_height * 72 / (dpi_y * max_lines * 1.18))
        starting_points = 10 if not fit_height else min(10, max(4, height_points))
        best_regular_font = None
        best_bold_font = None
        best_line_height = 1
        for points in range(starting_points, 3, -1):
            regular_font = self._create_print_font(gdi32, dpi_y, points)
            bold_font = self._create_print_font(gdi32, dpi_y, points, weight=700)
            size = wintypes.SIZE()
            widest = 0
            for line, bold in print_lines:
                old_font = gdi32.SelectObject(hdc, bold_font if bold else regular_font)
                gdi32.GetTextExtentPoint32W(hdc, line, len(line), ctypes.byref(size))
                gdi32.SelectObject(hdc, old_font)
                widest = max(widest, size.cx)

            line_height = max(1, int(points * dpi_y / 72 * 1.22))
            height_fits = (not fit_height) or line_height * max_lines <= printable_height
            if (not fit_width or widest <= printable_width) and height_fits:
                return regular_font, bold_font, line_height
            if best_regular_font:
                gdi32.DeleteObject(best_regular_font)
            if best_bold_font:
                gdi32.DeleteObject(best_bold_font)
            best_regular_font = regular_font
            best_bold_font = bold_font
            best_line_height = line_height
        return best_regular_font, best_bold_font, best_line_height

    def _create_print_font(self, gdi32, dpi_y, points, weight=400):
        height = -max(1, int(points * dpi_y / 72))
        return gdi32.CreateFontW(height, 0, 0, 0, weight, 0, 0, 0, 1, 0, 0, 5, 49, "Consolas")

    def _clip_line_for_print_width(self, gdi32, hdc, line, printable_width):
        size = wintypes.SIZE()
        gdi32.GetTextExtentPoint32W(hdc, line, len(line), ctypes.byref(size))
        if size.cx <= printable_width:
            return line
        if printable_width <= 0:
            return ""

        ellipsis = "..."
        low = 0
        high = max(0, len(line) - len(ellipsis))
        best = ellipsis
        while low <= high:
            mid = (low + high) // 2
            candidate = line[:mid].rstrip() + ellipsis
            gdi32.GetTextExtentPoint32W(hdc, candidate, len(candidate), ctypes.byref(size))
            if size.cx <= printable_width:
                best = candidate
                low = mid + 1
            else:
                high = mid - 1
        return best

    def _default_windows_printer(self):
        winspool = ctypes.WinDLL("winspool.drv", use_last_error=True)
        winspool.GetDefaultPrinterW.argtypes = [wintypes.LPWSTR, ctypes.POINTER(wintypes.DWORD)]
        winspool.GetDefaultPrinterW.restype = wintypes.BOOL
        needed = wintypes.DWORD(0)
        winspool.GetDefaultPrinterW(None, ctypes.byref(needed))
        if needed.value == 0:
            raise OSError("Nenhuma impressora padrão configurada no Windows.")

        buffer = ctypes.create_unicode_buffer(needed.value)
        if not winspool.GetDefaultPrinterW(buffer, ctypes.byref(needed)):
            raise ctypes.WinError(ctypes.get_last_error())
        return buffer.value

    def _format_print_workbook(self, sheet):
        header_fill = PatternFill("solid", fgColor=COLORS["table_header"].replace("#", ""))
        header_font = Font(color=COLORS["table_header_text"].replace("#", ""), bold=True)
        text_font = Font(color=COLORS["text"].replace("#", ""))

        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS) + 1)
        sheet.cell(row=1, column=1).font = Font(color=COLORS["text"].replace("#", ""), bold=True, size=13)
        sheet.cell(row=2, column=1).font = text_font
        sheet.cell(row=3, column=1).font = text_font

        header_row = 5
        for cell in sheet[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row in sheet.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.font = text_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        widths = [7, 12, 11, 22, 18, 12]
        for column_index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(column_index)].width = width

        sheet.freeze_panes = "A6"
        sheet.auto_filter.ref = f"A5:{get_column_letter(len(COLUMNS) + 1)}{sheet.max_row}"
        sheet.print_area = f"A1:{get_column_letter(len(COLUMNS) + 1)}{sheet.max_row}"
        sheet.page_setup.orientation = "portrait"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 1
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_margins.left = 0.25
        sheet.page_margins.right = 0.25
        sheet.page_margins.top = 0.35
        sheet.page_margins.bottom = 0.35

    def _send_file_to_printer(self, path):
        if sys.platform.startswith("win"):
            os.startfile(str(path), "print")
        elif sys.platform == "darwin":
            subprocess.run(["lp", str(path)], check=True)
        else:
            subprocess.run(["lp", str(path)], check=True)

    def print_report(self):
        self._finish_active_cell()

        report_lines = self._printable_report_lines()
        report = self._printable_report()
        try:
            if sys.platform.startswith("win"):
                report_lines = self._wrap_report_lines_for_print(report_lines, columns=1)
                self._send_print_lines_to_windows_printer(report_lines, "Resumo", columns=1, paginate=True)
            else:
                self._send_text_to_printer(report, "Resumo")
        except Exception as exc:
            path = Path(tempfile.gettempdir()) / f"iomarques_live_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            try:
                path.write_text(report, encoding="utf-8")
                saved_message = f"Salvei o relatório em:\n{path}\n\n"
            except OSError:
                saved_message = ""
            messagebox.showerror(
                "Não consegui imprimir",
                f"{saved_message}Erro ao enviar para a impressora:\n{exc}",
            )
            return

        messagebox.showinfo("Impressão enviada", "Enviei o resumo para a impressora padrão.")

    def clear_all(self):
        self._finish_active_cell()
        answer = messagebox.askyesno(
            "Começar nova live?",
            "Isso apaga todos os dados da tabela atual, o cronômetro e o salvamento automático. Deseja limpar tudo?",
        )
        if not answer:
            return

        self._save_rows(force_snapshot=True)

        if self.timer_job is not None:
            self.after_cancel(self.timer_job)
            self.timer_job = None
        self.live_running = False
        self.live_id = None
        self.live_first_started_at = None
        self.live_started_at = None
        self.live_finished_at = None
        self.live_elapsed_seconds = 0

        for child in self.body_frame.winfo_children():
            child.destroy()
        self.row_vars.clear()
        self.index_frames.clear()
        self.index_labels.clear()
        self.cell_frames.clear()
        self.cell_entries.clear()
        self.clear_buttons.clear()
        self.active_cell = None
        self.hovered_cell = None
        self._add_row()
        self._refresh_totals()
        self._refresh_live_controls()
        self._save_rows()

    def _finish_active_cell(self):
        if self.active_cell and self._cell_exists(*self.active_cell):
            self._finish_cell_edit(*self.active_cell)

    def _on_close(self):
        self._finish_active_cell()
        self._save_rows(force_snapshot=True)
        if self.timer_job is not None:
            self.after_cancel(self.timer_job)
        if self.autosave_job is not None:
            self.after_cancel(self.autosave_job)
        self.destroy()


if __name__ == "__main__":
    app = LiveSalesApp()
    app.mainloop()
